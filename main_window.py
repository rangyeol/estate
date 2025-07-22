from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QTabWidget, QTableWidget, QTableWidgetItem,
    QComboBox, QFrame, QMessageBox, QCheckBox, QRadioButton, QButtonGroup,
    QFileDialog, QProgressBar, QSplitter, QStatusBar, QStyle, QMenu, QToolBar, QProgressDialog
)
from PySide6.QtCore import Qt, Signal, Slot, QSize, QMetaObject, QDateTime, QTimer, QObject, Q_ARG
from PySide6.QtGui import QPixmap, QIcon, QAction, QFont, QCursor
from property_table import PropertyTable
from property_detail import PropertyDetailWidget
from naver_api import NaverLandAPI
from loading_dialog import LoadingDialog
import re
import os
import json
import pandas as pd
from datetime import datetime
import time
import copy
from openpyxl.styles import PatternFill # 엑셀 스타일링을 위해 추가

class ProgressSignal(QObject):
    progress_updated = Signal(int, str)
    search_completed = Signal(dict)
    search_failed = Signal(str)
    articles_found = Signal(object, list, int)
    show_warning = Signal(int, int, str)

class MainWindow(QMainWindow):
    def __init__(self, data=None):
        super().__init__()

        self.api = NaverLandAPI()
        self.data = data or {}
        self.articles = [] 
        self.original_articles = [] 
        self.complex_articles = []  
        self.original_complex_articles = []
        
        self.average_prices = {} # 그룹별 평균가 저장
        self.overall_average_prices = {} # 전체 평균가 저장

        self.search_history = []
        self.current_search_keyword = ""

        self.progress_signal = ProgressSignal()
        self.progress_signal.progress_updated.connect(self.update_progress)
        self.progress_signal.search_completed.connect(self.on_search_completed)
        self.progress_signal.search_failed.connect(self.on_search_failed)
        self.progress_signal.articles_found.connect(self.update_article_table)
        self.progress_signal.show_warning.connect(self.show_warning_dialog)

        if data:
            if 'complexes' in data:
                self.articles = data.get('complexes', [])
            elif 'articleList' in data: 
                self.complex_articles = data.get('articleList', [])
                self.original_complex_articles = self.complex_articles.copy()
            if self.articles:
                self.original_articles = self.articles.copy()
            else:
                self.original_articles = []

        import threading
        self.user_choice_event = threading.Event()
        self.collect_all_pages = False
        
        self.detail_window_instance = None 
        self.loading_dialog = None

        self.apply_modern_style()
        self.init_ui()
        
        self.desired_excel_columns_korean = [
            '매물명', '부동산유형', '매물부동산유형', '거래유형', '확인유형', '층정보', '가격변동상태',
            '매매/전세가', '월세', '면적명', '전용면적', '공급면적', '방향', '매물확인일',
            '매물특징', '태그목록', '동이름', '동일주소매물수', '동일주소최고가', '동일주소최저가',
            '위도', '경도', '부동산중개사명', '소유자확인거래', '직거래여부', '관심여부',
            '단지여부', '상세주소', '안심임대인여부', '단지명', '단지번호', '주소',
            '지역코드', '링크', '최고층', '최저층', '최대공급면적', '최대전체면적',
            '최소공급면적', '최소전체면적', '총동수', '총세대수', '사용승인일', '사용여부'
        ]

    def apply_modern_style(self):
        primary_color = "#0077b6"; secondary_color = "#00b4d8"; accent_color = "#48cae4"
        success_color = "#2dc653"; warning_color = "#ffaa00"; error_color = "#d62828"
        text_color = "#1d3557"; light_text_color = "#ffffff"; background_color = "#f8f9fa"
        card_color = "#ffffff"; border_color = "#e9ecef"; hover_color = "#e0e5ec"
        self.setStyleSheet(f"""
            QMainWindow, QDialog {{ background-color: {background_color}; }}
            QWidget {{ font-family: 'Segoe UI', 'Malgun Gothic', sans-serif; font-size: 10pt; color: {text_color}; }}
            QLabel {{ padding: 2px; color: {text_color}; font-weight: normal; }}
            #summaryLabel {{ font-weight: bold; color: {primary_color}; }}
            #priceLabel {{ color: #2c3e50; font-size: 9.5pt; padding: 0 10px; }}
            #summaryFrame {{ background-color: #e9ecef; border-radius: 5px; }}
            QLabel[objectName^="header"] {{ font-weight: bold; font-size: 11pt; color: {primary_color}; margin-bottom: 4px; }}
            QPushButton {{ background-color: {primary_color}; color: {light_text_color}; border: none; border-radius: 4px; padding: 6px 10px; font-weight: bold; min-height: 18px; }}
            QPushButton:hover {{ background-color: {secondary_color}; }}
            QPushButton:pressed {{ background-color: {accent_color}; }}
            QPushButton:disabled {{ background-color: #c8d6e5; color: #8395a7; }}
            QLineEdit {{ border: 1px solid {border_color}; border-radius: 4px; padding: 6px; background-color: white; selection-background-color: {primary_color}; }}
            QLineEdit:focus {{ border: 1px solid {primary_color}; }}
            QTableWidget {{ gridline-color: {border_color}; background-color: {card_color}; border: 1px solid {border_color}; border-radius: 4px; selection-background-color: {accent_color}; selection-color: {text_color}; alternate-background-color: #f5f9fc; }}
            QTableWidget::item {{ padding: 4px; border: none; }}
            QTableWidget::item:selected {{ background-color: {accent_color}; color: {light_text_color}; }}
            QHeaderView::section {{ background-color: {primary_color}; color: {light_text_color}; padding: 5px; border: none; font-weight: bold; }}
            QHeaderView::section:checked {{ background-color: {secondary_color}; }}
            QSplitter::handle {{ background-color: {border_color}; height: 1px; }}
            QStatusBar {{ background-color: {primary_color}; color: {light_text_color}; padding: 3px; font-size: 9pt; }}
        """)
        self.search_button_style = f"QPushButton {{ background-color: {primary_color}; color: {light_text_color}; border: none; border-radius: 4px; padding: 6px 12px; font-weight: bold; min-height: 20px; }} QPushButton:hover {{ background-color: {secondary_color}; }}"
        self.primary_button_style = f"QPushButton {{ background-color: {success_color}; color: {light_text_color}; border: none; border-radius: 4px; padding: 6px 12px; font-weight: bold; min-height: 20px; }} QPushButton:hover {{ background-color: #26a64a; }} QPushButton:disabled {{ background-color: #a5d6a7; color: #e8f5e9; }}"
        self.secondary_button_style = f"QPushButton {{ background-color: #f5f9fc; color: {primary_color}; border: 1px solid {primary_color}; border-radius: 4px; padding: 5px 10px; font-weight: bold; min-height: 20px; }} QPushButton:hover {{ background-color: #e6f2ff; color: {secondary_color}; border-color: {secondary_color}; }}"

    def init_ui(self):
        self.setWindowTitle("네이버 부동산 데이터 뷰어")
        self.resize(1280, 800)
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10); main_layout.setSpacing(10)
        
        search_layout = QHBoxLayout()
        search_label = QLabel("검색어:"); search_label.setFixedWidth(50); search_layout.addWidget(search_label)
        self.search_input = QLineEdit(); self.search_input.setPlaceholderText("지역명 또는 아파트명을 입력하세요"); self.search_input.returnPressed.connect(self.fetch_from_api); search_layout.addWidget(self.search_input)
        self.search_button = QPushButton("검색"); self.search_button.setStyleSheet(self.search_button_style) 
        self.search_button.clicked.connect(self.fetch_from_api); search_layout.addWidget(self.search_button)
        main_layout.addLayout(search_layout)
        
        self.splitter = QSplitter(Qt.Orientation.Vertical); main_layout.addWidget(self.splitter, 1)
        
        complex_frame = QWidget(); complex_layout = QVBoxLayout(complex_frame); complex_layout.setContentsMargins(0,0,0,0)
        complex_header_layout = QHBoxLayout(); complex_header_layout.setContentsMargins(0,0,0,0)
        complex_header = QLabel("단지 목록"); complex_header.setObjectName("header"); complex_header.setFixedHeight(30); complex_header_layout.addWidget(complex_header)
        self.complex_type_group = QButtonGroup(self); self.complex_filter_layout = QHBoxLayout(); self.complex_filter_layout.setContentsMargins(0,0,0,0); self.complex_filter_layout.setSpacing(5)
        complex_header_layout.addLayout(self.complex_filter_layout); complex_header_layout.addStretch()
        self.search_checked_button = QPushButton("선택 단지 검색"); self.search_checked_button.setStyleSheet(self.primary_button_style); self.search_checked_button.clicked.connect(self.search_checked_complexes); complex_header_layout.addWidget(self.search_checked_button)
        self.complex_select_all_button = QPushButton("전체선택"); self.complex_select_all_button.setStyleSheet(self.secondary_button_style); self.complex_select_all_button.clicked.connect(lambda: self.complex_table.check_all_items(True)); complex_header_layout.addWidget(self.complex_select_all_button)
        self.complex_deselect_all_button = QPushButton("모두해제"); self.complex_deselect_all_button.setStyleSheet(self.secondary_button_style); self.complex_deselect_all_button.clicked.connect(lambda: self.complex_table.check_all_items(False)); complex_header_layout.addWidget(self.complex_deselect_all_button)
        complex_layout.addLayout(complex_header_layout)
        self.complex_table = PropertyTable(is_complex_table=True); self.complex_table.checkbox_toggled.connect(self.on_complex_checkbox_toggled)
        complex_layout.addWidget(self.complex_table); self.splitter.addWidget(complex_frame)
        
        property_frame = QWidget(); property_layout = QVBoxLayout(property_frame); property_layout.setContentsMargins(0,0,0,0)
        property_header_layout = QHBoxLayout(); property_header_layout.setContentsMargins(0,0,0,0)
        property_header = QLabel("매물 목록"); property_header.setObjectName("header"); property_header.setFixedHeight(30); property_header_layout.addWidget(property_header)
        self.article_type_group = QButtonGroup(self); self.article_filter_layout = QHBoxLayout(); self.article_filter_layout.setContentsMargins(0,0,0,0); self.article_filter_layout.setSpacing(5)
        property_header_layout.addLayout(self.article_filter_layout); property_header_layout.addStretch()
        self.download_button = QPushButton("엑셀 다운로드"); self.download_button.setStyleSheet(self.primary_button_style); self.download_button.clicked.connect(self.download_to_excel); property_header_layout.addWidget(self.download_button)
        self.article_select_all_button = QPushButton("전체선택"); self.article_select_all_button.setStyleSheet(self.secondary_button_style); self.article_select_all_button.clicked.connect(lambda: self.property_table.check_all_items(True)); property_header_layout.addWidget(self.article_select_all_button)
        self.article_deselect_all_button = QPushButton("모두해제"); self.article_deselect_all_button.setStyleSheet(self.secondary_button_style); self.article_deselect_all_button.clicked.connect(lambda: self.property_table.check_all_items(False)); property_header_layout.addWidget(self.article_deselect_all_button)
        property_layout.addLayout(property_header_layout)
        self.property_table = PropertyTable(self, is_complex_table=False); self.property_table.checkbox_toggled.connect(self.on_article_checkbox_toggled); self.property_table.detail_button_clicked.connect(self.on_article_detail_clicked)
        property_layout.addWidget(self.property_table); self.splitter.addWidget(property_frame)
        
        self.splitter.setSizes([int(self.height() * 0.4), int(self.height() * 0.6)])

        self.summary_widget = QFrame()
        self.summary_widget.setObjectName("summaryFrame")
        self.summary_layout = QHBoxLayout(self.summary_widget)
        self.summary_layout.setContentsMargins(15, 5, 15, 5)
        summary_title = QLabel("<b>거래 유형별 평균가</b>")
        summary_title.setObjectName("summaryLabel")
        self.avg_sale_label = QLabel("매매: -")
        self.avg_sale_label.setObjectName("priceLabel")
        self.avg_lease_label = QLabel("전세: -")
        self.avg_lease_label.setObjectName("priceLabel")
        self.summary_layout.addStretch()
        self.summary_layout.addWidget(summary_title)
        self.summary_layout.addWidget(self.avg_sale_label)
        self.summary_layout.addWidget(self.avg_lease_label)
        self.summary_layout.addStretch()
        self.summary_widget.setVisible(False)
        main_layout.addWidget(self.summary_widget)

        self.setStatusBar(QStatusBar()); self.statusBar().showMessage("준비 완료")
        self.progress_bar = QProgressBar(); self.progress_bar.setRange(0,100); self.progress_bar.setTextVisible(True); self.progress_bar.setVisible(False); self.statusBar().addPermanentWidget(self.progress_bar)
        self.create_menus(); self.create_toolbar()

        self._update_filter_radios(self.complex_filter_layout, self.complex_type_group, self.original_articles, 'realEstateTypeName', self.filter_complexes_by_type)
        self._update_filter_radios(self.article_filter_layout, self.article_type_group, self.original_complex_articles, 'tradeTypeName', self.filter_articles_by_trade_type)
        if self.articles: self.complex_table.load_data(self.articles)
        if self.complex_articles: self.property_table.load_data(self.complex_articles)
    
    def _format_price_to_korean_won(self, price_in_won):
        if not isinstance(price_in_won, (int, float)) or price_in_won <= 0:
            return "-"
        eok = int(price_in_won // 100000000)
        man = int((price_in_won % 100000000) // 10000)
        
        result = ""
        if eok > 0:
            result += f"{eok}억"
        if man > 0:
            result += f" {man:,}만원"
        
        return result.strip() if result else "-"

    def _update_average_price_display(self):
        if not self.overall_average_prices:
            self.summary_widget.setVisible(False)
            return

        sale_avg = self.overall_average_prices.get('매매')
        lease_avg = self.overall_average_prices.get('전세')

        self.avg_sale_label.setText(f"매매: {self._format_price_to_korean_won(sale_avg) if sale_avg else '-'}")
        self.avg_lease_label.setText(f"전세: {self._format_price_to_korean_won(lease_avg) if lease_avg else '-'}")
        
        self.summary_widget.setVisible(True)

    def _get_area_key(self, article):
        supply_area_str = article.get('area2', '')
        exclusive_area_str = article.get('area1', '')
        
        def _format_num(num_str):
            if not num_str: return ''
            try: return f"{float(num_str):.2f}".rstrip('0').rstrip('.')
            except (ValueError, TypeError): return str(num_str)

        s_area_fmt = _format_num(supply_area_str)
        e_area_fmt = _format_num(exclusive_area_str)

        if s_area_fmt and e_area_fmt: return f"{s_area_fmt}㎡ ({e_area_fmt}㎡)"
        elif s_area_fmt: return f"{s_area_fmt}㎡"
        elif e_area_fmt: return f"{e_area_fmt}㎡"
        return "-"
        
    def _calculate_and_store_average_prices(self, articles):
        if not articles:
            self.average_prices = {}
            self.overall_average_prices = {}
            return

        price_groups = {}
        overall_prices = {'매매': [], '전세': [], '월세': []}

        for article in articles:
            area_key = self._get_area_key(article)
            trade_type = article.get('tradeTypeName', '')
            
            key = (
                article.get('articleName', article.get('complexName', '')),
                trade_type,
                area_key
            )
            price = self._get_excel_numeric_price(article.get('dealOrWarrantPrc'))

            if price is not None and price > 0:
                if key not in price_groups: price_groups[key] = []
                price_groups[key].append(price)
                
                if trade_type in overall_prices:
                    overall_prices[trade_type].append(price)
        
        self.average_prices = {key: sum(prices) / len(prices) for key, prices in price_groups.items() if prices}
        
        self.overall_average_prices = {}
        for trade_type, prices in overall_prices.items():
            if prices:
                self.overall_average_prices[trade_type] = sum(prices) / len(prices)
        
        self._update_average_price_display()

    @Slot(object, list, int)
    def update_article_table(self, articles, complex_names_found, original_checked_count):
        try:
            if self.loading_dialog: self.loading_dialog.close()
            self.hide_progress_bar()
            self.restore_search_checked_button()
            self.complex_articles = articles if articles else []
            self.original_complex_articles = self.complex_articles.copy()

            if not self.complex_articles:
                self.property_table.load_data([])
                if hasattr(self, 'article_label'): self.article_label.setText("선택된 단지의 매물이 없습니다.")
                self.statusBar().showMessage("선택한 단지에서 매물을 찾지 못했습니다.")
                self._update_filter_radios(self.article_filter_layout, self.article_type_group, [], 'tradeTypeName', self.filter_articles_by_trade_type)
                self.summary_widget.setVisible(False)
                return

            self._calculate_and_store_average_prices(self.complex_articles)
            self.property_table.load_data(self.complex_articles, self.average_prices)
            
            total_articles_found = len(self.complex_articles)
            num_complexes_with_articles = len(complex_names_found)
            display_name_str = f"{original_checked_count}개 단지"
            if num_complexes_with_articles > 0:
                display_name_str = f"{complex_names_found[0]} 등 {num_complexes_with_articles}곳" if num_complexes_with_articles > 1 else complex_names_found[0]
            
            if hasattr(self, 'article_label'): self.article_label.setText(f"{display_name_str} 매물: {total_articles_found}건")
            self._update_filter_radios(self.article_filter_layout, self.article_type_group, self.original_complex_articles, 'tradeTypeName', self.filter_articles_by_trade_type)
            if hasattr(self, 'download_button'):
                self.download_button.setEnabled(True); self.download_button.setText(f"엑셀 다운로드 ({total_articles_found}건)"); self.download_button.setStyleSheet(self.primary_button_style)
            self.statusBar().showMessage(f"{display_name_str}에서 {total_articles_found}개의 매물을 찾았습니다.")
        except Exception as e: 
            import traceback
            traceback.print_exc()
            self.statusBar().showMessage(f"매물 표시 오류: {e}")

    def download_to_excel(self):
        articles_to_process = []
        is_selection_download = False
        
        if hasattr(self, 'complex_articles') and self.complex_articles:
            checked_article_indices = self.property_table.get_checked_items()
            if checked_article_indices:
                articles_to_process = [self.complex_articles[i] for i in checked_article_indices if 0 <= i < len(self.complex_articles)]
                is_selection_download = True
            else:
                articles_to_process = self.complex_articles
        else:
            self.statusBar().showMessage("다운로드할 매물 데이터가 없습니다."); return

        if not articles_to_process: self.statusBar().showMessage("다운로드할 데이터가 없습니다."); return

        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"매물목록_{len(articles_to_process)}건_{now}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "엑셀 파일 저장", default_filename, "Excel Files (*.xlsx)")
        if not file_path: return

        self.statusBar().showMessage(f"엑셀 파일 생성 중...")
        
        # 엑셀 데이터 생성 및 평균가 비교를 위한 준비
        excel_data_rows = []
        # 그룹별 평균가는 self.average_prices 에 이미 계산되어 있음
        
        for article_data in articles_to_process:
            # (기존 로직과 동일하게 엑셀 행 데이터 생성)
            excel_row = self._create_excel_row(article_data)
            excel_data_rows.append(excel_row)

        df = pd.DataFrame(excel_data_rows)
        if not self.desired_excel_columns_korean in df.columns:
             df = df.reindex(columns=self.desired_excel_columns_korean)

        try:
            # 스타일 적용을 위해 ExcelWriter 사용
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='매물목록')
                worksheet = writer.sheets['매물목록']

                # 색상 채우기 스타일 정의
                sale_fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid') # 선홍색
                lease_fill = PatternFill(start_color='E9F5E9', end_color='E9F5E9', fill_type='solid') # 연녹색
                rent_fill = PatternFill(start_color='E0F7FA', end_color='E0F7FA', fill_type='solid') # 하늘색

                # 각 행을 순회하며 조건부 서식 적용
                for index, article in enumerate(articles_to_process):
                    excel_row_num = index + 2  # 엑셀은 1-based, 헤더 포함
                    
                    key = (
                        article.get('articleName', article.get('complexName', '')),
                        article.get('tradeTypeName', ''),
                        self._get_area_key(article)
                    )
                    avg_price = self.average_prices.get(key)
                    price = self._get_excel_numeric_price(article.get('dealOrWarrantPrc'))
                    
                    if avg_price and price and price < avg_price:
                        trade_type = article.get('tradeTypeName', '')
                        fill_to_apply = None
                        if trade_type == '매매': fill_to_apply = sale_fill
                        elif trade_type == '전세': fill_to_apply = lease_fill
                        elif trade_type == '월세': fill_to_apply = rent_fill
                        
                        if fill_to_apply:
                            for col_num in range(1, len(df.columns) + 1):
                                worksheet.cell(row=excel_row_num, column=col_num).fill = fill_to_apply
                
                # 컬럼 너비 자동 조정
                for column_cells in worksheet.columns:
                    try:
                        max_length = 0
                        column = column_cells[0].column_letter
                        for cell in column_cells:
                            if cell.value:
                                # 한글은 글자당 2, 영문/숫자는 1로 계산하여 너비 추정
                                length = sum(2 if '\uac00' <= char <= '\ud7a3' else 1 for char in str(cell.value))
                                if length > max_length:
                                    max_length = length
                        adjusted_width = max_length + 2
                        worksheet.column_dimensions[column].width = adjusted_width
                    except:
                        pass

            self.statusBar().showMessage(f"엑셀 파일 저장 완료: {file_path} ({len(df)}건)")
        except Exception as e:
            self.statusBar().showMessage(f"엑셀 파일 저장 오류: {e}")
            QMessageBox.critical(self, "엑셀 저장 오류", f"엑셀 파일 저장 중 오류가 발생했습니다:\n{str(e)}")
            import traceback; traceback.print_exc()

    def _create_excel_row(self, article_data):
        # 엑셀 행 생성을 위한 헬퍼 함수
        excel_column_api_map = {
            '매물명': 'articleName', '부동산유형': 'realEstateTypeName', '매물부동산유형': 'articleRealEstateTypeName',
            '거래유형': 'tradeTypeName', '확인유형': 'verificationTypeCode', '층정보': 'floorInfo', '가격변동상태': 'priceChangeState',
            '매매/전세가': 'dealOrWarrantPrc', '월세': 'rentPrc', '면적명': 'areaName', '전용면적': 'area1', '공급면적': 'area2',
            '방향': 'direction', '매물확인일': 'articleConfirmYmd', '매물특징': 'articleFeatureDesc', '태그목록': 'tagList',
            '동이름': 'buildingName', '동일주소매물수': 'sameAddrCnt', '동일주소최고가': 'sameAddrMaxPrc', '동일주소최저가': 'sameAddrMinPrc',
            '위도': 'latitude', '경도': 'longitude', '부동산중개사명': 'realtorName', '소유자확인거래': 'tradeCheckedByOwner',
            '직거래여부': 'isDirectTrade', '관심여부': 'isInterest', '단지여부': 'isComplex', '상세주소': 'detailAddress',
            '안심임대인여부': 'isSafeLessorOfHug', '단지명': 'complexName', '단지번호': 'complexNo', '주소': 'cortarAddressFromComplex',
            '지역코드': 'cortarNoFromComplex', '링크': 'cpPcArticleUrl', '최고층': 'highFloorFromComplex', '최저층': 'lowFloorFromComplex',
            '최대공급면적': 'maxSupplyAreaFromComplex', '최대전체면적': 'maxTotalAreaFromComplex', '최소공급면적': 'minSupplyAreaFromComplex',
            '최소전체면적': 'minTotalAreaFromComplex', '총동수': 'totalDongCountFromComplex', '총세대수': 'totalHouseholdCountFromComplex',
            '사용승인일': 'useApproveYmdFromComplex', '사용여부': 'useYnFromComplex'
        }
        verification_mapping = {'NONE': '없음', 'OWNER': '소유자확인', 'REALTOR': '중개사확인', 'LESSOR': '임대인확인', 'S_VR': 'VR확인', 'SITE': '현장확인', 'NDOC1': '서류확인1', 'DOC': '서류확인', 'DOCV2': '서류확인V2', 'MOBL': '모바일확인', 'NDOC2': '서류확인2'}
        price_change_mapping = {'SAME': '변동없음', 'DOWN': '하락', 'UP': '상승', 'DECREASE': '감소', 'INCREASE': '증가', 'NEW': '신규'}

        excel_row = {}
        for korean_col_name in self.desired_excel_columns_korean:
            api_key = excel_column_api_map.get(korean_col_name)
            raw_value = article_data.get(api_key, '-') if api_key else '-'

            if korean_col_name == '매매/전세가': excel_row[korean_col_name] = self._get_excel_numeric_price(article_data.get('dealOrWarrantPrc'))
            elif korean_col_name == '월세': excel_row[korean_col_name] = self._get_excel_numeric_rent(article_data.get('rentPrc'))
            elif korean_col_name in ['매물확인일', '사용승인일']: excel_row[korean_col_name] = self._format_excel_date(raw_value if isinstance(raw_value, str) else str(raw_value))
            elif korean_col_name == '태그목록': excel_row[korean_col_name] = ', '.join(raw_value) if isinstance(raw_value, list) else raw_value
            elif korean_col_name == '확인유형': excel_row[korean_col_name] = verification_mapping.get(str(raw_value).upper(), raw_value)
            elif korean_col_name == '가격변동상태': excel_row[korean_col_name] = price_change_mapping.get(str(raw_value).upper(), raw_value)
            elif korean_col_name in ['소유자확인거래', '직거래여부', '관심여부', '단지여부', '안심임대인여부', '사용여부']: excel_row[korean_col_name] = 'Y' if raw_value is True or str(raw_value).upper() == 'Y' else ('N' if raw_value is False or str(raw_value).upper() == 'N' else raw_value)
            elif korean_col_name == '주소':
                full_addr_parts = [p for p in [article_data.get('cortarAddressFromComplex', ''), article_data.get('buildingName', ''), article_data.get('detailAddress', '')] if p]
                excel_row[korean_col_name] = " ".join(full_addr_parts).strip() or '-'
            elif korean_col_name == '매물명': excel_row[korean_col_name] = article_data.get('articleName', article_data.get('complexName', '-'))
            else: excel_row[korean_col_name] = raw_value if raw_value is not None else "-"
        return excel_row
        
    @Slot()
    def filter_articles_by_trade_type(self, button=None):
        if not button or not hasattr(self, 'original_complex_articles'): 
            if hasattr(self, 'property_table'): self.property_table.load_data([])
            if hasattr(self, 'article_label'): self.article_label.setText("매물 목록: 0건")
            self._update_filter_radios(self.article_filter_layout, self.article_type_group, [], 'tradeTypeName', self.filter_articles_by_trade_type)
            self.summary_widget.setVisible(False)
            return

        selected_type = button.text()
        if selected_type == "전체": self.complex_articles = self.original_complex_articles.copy()
        else: self.complex_articles = [a for a in self.original_complex_articles if a.get('tradeTypeName') == selected_type]
        
        self._calculate_and_store_average_prices(self.complex_articles)
        self.property_table.load_data(self.complex_articles, self.average_prices)

        total_articles = len(self.complex_articles)
        if hasattr(self, 'download_button'): self.download_button.setText(f"엑셀 다운로드 ({total_articles}건)")
        self.on_article_checkbox_toggled(-1, False)

    # ... (기존의 다른 메서드들은 변경 없이 유지)
    # fetch_from_api, on_search_completed, on_search_failed, hide_progress_bar, hide_loading_dialog,
    # add_to_history, search_checked_complexes, search_multiple_complexes, on_complex_checkbox_toggled,
    # restore_search_checked_button, show_warning_dialog, on_article_detail_clicked, show_property_detail_from_data,
    # on_article_checkbox_toggled, enable_search_button, create_menus, create_toolbar, 
    # _get_excel_numeric_price, _get_excel_numeric_rent, _format_excel_date, filter_complexes_by_type
    
    # ... (생략된 기존 코드는 이전 답변과 동일)
    
    # 이전에 제공된 전체 코드를 여기에 붙여넣으시면 됩니다. 
    # 이 답변에서는 새로 추가되거나 핵심적으로 수정된 부분 위주로 보여드렸습니다.
    # 아래는 생략된 부분의 시작입니다.
    def _update_filter_radios(self, layout, button_group, data, key, callback):
        for btn in button_group.buttons(): button_group.removeButton(btn); layout.removeWidget(btn); btn.deleteLater()
        all_radio = QRadioButton("전체"); all_radio.setChecked(True); button_group.addButton(all_radio); layout.addWidget(all_radio)
        all_radio.toggled.connect(lambda checked: callback(all_radio) if checked and callable(callback) else None)
        if all_radio.isChecked() and callable(callback): QTimer.singleShot(0, lambda: callback(all_radio))
        if data:
            unique_values = sorted(list(set(d.get(key, '') for d in data if d.get(key))))
            for value in unique_values:
                radio = QRadioButton(value); button_group.addButton(radio); layout.addWidget(radio)
                radio.toggled.connect(lambda checked, b=radio: callback(b) if checked and callable(callback) else None)

    
    @Slot()
    def fetch_from_api(self):
        self.statusBar().showMessage("API에서 데이터 가져오는 중...")
        search_input = self.search_input.text().strip()
        if not search_input: 
            self.statusBar().showMessage("검색어를 입력해주세요.")
            return

        if not self.loading_dialog:
            self.loading_dialog = LoadingDialog(self)
        self.loading_dialog.show()

        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("검색 준비 중...")
        self.search_button.setEnabled(False)
        keyword = search_input
        self.current_search_keyword = f"키워드: {keyword}"
        import threading
        def search_thread():
            try:
                self.progress_signal.progress_updated.emit(10, f"키워드 '{keyword}'로 검색 시작...")
                self.api.set_progress_callback(self.update_search_progress)
                result = self.api.search_by_keyword(keyword, max_pages=None)
                if result: self.progress_signal.search_completed.emit(result)
                else: self.progress_signal.search_failed.emit("API 응답을 가져올 수 없습니다.")
            except Exception as e: self.progress_signal.search_failed.emit(f"검색 중 오류 발생: {str(e)}")
        thread = threading.Thread(target=search_thread)
        thread.daemon = True
        thread.start()

    def update_search_progress(self, progress, message, status=None):
        self.progress_signal.progress_updated.emit(progress, message)

    @Slot(int, str)
    def update_progress(self, value, text):
        self.progress_bar.setValue(value)
        self.progress_bar.setFormat(f"{value}% - {text}")
        self.statusBar().showMessage(text)
        
    @Slot(dict)
    def on_search_completed(self, result):
        if self.loading_dialog:
            self.loading_dialog.close()

        self.progress_bar.setValue(100)
        self.progress_bar.setFormat("검색 완료!")
        self.search_button.setEnabled(True)
        self.data = result
        if 'error' in self.data:
            self.statusBar().showMessage(f"API 오류: {self.data.get('error', '알 수 없는 오류')}")
            self.hide_progress_bar()
            return
        
        complexes_data = self.data.get('complexes', self.data.get('list', []))
        self.articles = complexes_data
        if self.articles:
             self.original_articles = self.articles.copy()
        else:
             self.articles = []
             self.original_articles = []

        total_complexes = len(self.articles)

        if total_complexes > 0:
            msg = f"API에서 {total_complexes}개의 단지 정보를 가져왔습니다."
            self.add_to_history(f"[단지] {self.current_search_keyword} ({total_complexes}개)")
            self.complex_table.load_data(self.articles)
            self._update_filter_radios(self.complex_filter_layout, self.complex_type_group, self.original_articles, 'realEstateTypeName', self.filter_complexes_by_type)
            if hasattr(self, 'complex_label'): self.complex_label.setText(f"단지 목록: {total_complexes}개")
            self.property_table.load_data([])
            self.complex_articles = []
            self.original_complex_articles = []
            if hasattr(self, 'article_label'): self.article_label.setText("선택된 단지의 매물 목록: 0개")
            self._update_filter_radios(self.article_filter_layout, self.article_type_group, [], 'tradeTypeName', self.filter_articles_by_trade_type)
            self.statusBar().showMessage(msg)
            QMessageBox.information(self, "검색 완료", f"키워드 '{self.search_input.text().strip()}'로 {total_complexes}개의 단지를 찾았습니다.\n\n1. 단지 목록에서 단지를 선택(체크)하고 '선택 단지 검색'을 누르세요.\n2. 매물 목록에서 '상세보기'를 누르세요.")
        else: 
            self.statusBar().showMessage("검색 결과가 없습니다.")
            self.complex_table.load_data([])
            self._update_filter_radios(self.complex_filter_layout, self.complex_type_group, [], 'realEstateTypeName', self.filter_complexes_by_type)
        self.hide_progress_bar()

    @Slot(str)
    def on_search_failed(self, error_message):
        if self.loading_dialog:
            self.loading_dialog.close()
        self.hide_progress_bar()
        self.search_button.setEnabled(True)
        self.statusBar().showMessage(f"검색 실패: {error_message}")
        self.complex_table.load_data([])
        self.property_table.load_data([])
        self.articles = []
        self.original_articles = []
        self.complex_articles = []
        self.original_complex_articles = []
        if hasattr(self, 'complex_label'): self.complex_label.setText("단지 목록: 0개")
        if hasattr(self, 'article_label'): self.article_label.setText("선택된 단지의 매물 목록: 0개")
        self._update_filter_radios(self.complex_filter_layout, self.complex_type_group, [], 'realEstateTypeName', self.filter_complexes_by_type)
        self._update_filter_radios(self.article_filter_layout, self.article_type_group, [], 'tradeTypeName', self.filter_articles_by_trade_type)

    def hide_progress_bar(self):
        self.progress_bar.hide()

    @Slot()
    def hide_loading_dialog(self):
        if self.loading_dialog:
            self.loading_dialog.close()

    def add_to_history(self, label):
        history_item = {'label': label, 'data': copy.deepcopy(self.data), 
                        'articles': copy.deepcopy(self.articles), 
                        'complex_articles': copy.deepcopy(self.complex_articles),
                        'keyword': self.current_search_keyword}
        self.search_history.insert(0, history_item)
        if len(self.search_history) > 10: self.search_history = self.search_history[:10]

    @Slot(int, int)
    def on_complex_cell_clicked_improved(self, row, column):
        try:
            if column > 0: self.complex_table.selectRow(row)
        except Exception as e: import traceback; traceback.print_exc()

    def search_checked_complexes(self):
        checked_complex_data_list = self.complex_table.get_checked_data()
        if not checked_complex_data_list: 
            self.statusBar().showMessage("검색할 단지를 체크박스로 선택해주세요.")
            return
        complex_names = [c.get('complexName', '단지') for c in checked_complex_data_list]
        names_str = ", ".join(complex_names[:3]) + (f" 외 {len(complex_names)-3}개" if len(complex_names)>3 else "")
        self.statusBar().showMessage(f"체크된 단지({len(checked_complex_data_list)}개): {names_str} - 데이터 수집중...")
        if hasattr(self, 'article_label'): self.article_label.setText(f"선택된 단지의 매물 목록: {names_str}")
        self.property_table.load_data([]) 
        self.search_checked_button.setEnabled(False)
        self.search_checked_button.setText("검색 중...")
        self.search_multiple_complexes(checked_complex_data_list)

    @Slot()
    def on_complex_selected(self):
        if hasattr(self, 'progress_bar'): self.progress_bar.hide()

    def search_multiple_complexes(self, complexes_to_search):
        if not complexes_to_search: 
            self.restore_search_checked_button()
            return
        self.search_button.setEnabled(False)
        
        if not self.loading_dialog:
            self.loading_dialog = LoadingDialog(self)
        self.loading_dialog.show()

        import threading
        def search_thread_worker():
            all_found_articles, names_of_complexes_with_articles, failed_complex_searches = [], [], []
            try:
                self.api.set_progress_callback(self.update_search_progress)
                for i, complex_item in enumerate(complexes_to_search):
                    complex_no = complex_item.get('complexNo')
                    c_name = complex_item.get('complexName', 'N/A')
                    complex_level_data_to_inject = {
                        'complexName': c_name,
                        'complexNo': complex_no,
                        'cortarAddressFromComplex': complex_item.get('cortarAddress', ''),
                        'cortarNoFromComplex': complex_item.get('cortarNo', ''),
                        'totalDongCountFromComplex': complex_item.get('totalDongCount', ''),
                        'totalHouseholdCountFromComplex': complex_item.get('totalHouseholdCount', ''),
                        'useApproveYmdFromComplex': complex_item.get('useApproveYmd', ''),
                        'highFloorFromComplex': complex_item.get('highFloor', ''),
                        'lowFloorFromComplex': complex_item.get('lowFloor', ''),
                        'maxSupplyAreaFromComplex': complex_item.get('maxSupplyArea', ''),
                        'maxTotalAreaFromComplex': complex_item.get('maxTotalArea', ''),
                        'minSupplyAreaFromComplex': complex_item.get('minSupplyArea', ''),
                        'minTotalAreaFromComplex': complex_item.get('minTotalArea', ''),
                        'useYnFromComplex': complex_item.get('useYn', '')
                    }

                    if not complex_no: 
                        failed_complex_searches.append(f"{c_name}(번호없음)")
                        continue
                    
                    self.progress_signal.progress_updated.emit(int((i / len(complexes_to_search)) * 100), f"{c_name} 검색중 ({i+1}/{len(complexes_to_search)})")
                    try:
                        api_result = self.api.search_by_complex(complex_no=complex_no)
                        if api_result and isinstance(api_result, dict):
                            current_complex_articles = api_result.get('articleList', [])
                            if current_complex_articles:
                                for art in current_complex_articles: 
                                    art.update(complex_level_data_to_inject)
                                all_found_articles.extend(current_complex_articles)
                                if c_name not in names_of_complexes_with_articles: names_of_complexes_with_articles.append(c_name)
                            else: failed_complex_searches.append(f"{c_name}(매물X)")
                        else: failed_complex_searches.append(f"{c_name}(API오류)")
                    except Exception as e_inner: 
                        failed_complex_searches.append(f"{c_name}(에러:{str(e_inner)[:30]})")
                        print(f"개별 단지({c_name}) 매물 검색 오류: {e_inner}")
                
                if all_found_articles:
                    self.progress_signal.articles_found.emit(all_found_articles, names_of_complexes_with_articles, len(complexes_to_search))
                else:
                    self.progress_signal.search_failed.emit(f"선택 단지 매물 없음. 실패: {', '.join(failed_complex_searches)}")
            except Exception as e_outer: 
                self.progress_signal.search_failed.emit(f"매물 검색 중 전체 오류: {e_outer}")
                print(f"전체 매물 검색 오류: {e_outer}")
            finally: 
                QMetaObject.invokeMethod(self, "hide_loading_dialog", Qt.ConnectionType.QueuedConnection)
                QMetaObject.invokeMethod(self, "enable_search_button", Qt.ConnectionType.QueuedConnection, Q_ARG(str, "검색"))
                QMetaObject.invokeMethod(self, "restore_search_checked_button", Qt.ConnectionType.QueuedConnection)
        
        thread = threading.Thread(target=search_thread_worker)
        thread.daemon = True
        thread.start()

    @Slot(object, list, int)
    def update_article_table(self, articles, complex_names_found, original_checked_count):
        try:
            if self.loading_dialog:
                self.loading_dialog.close()
            self.hide_progress_bar()
            self.restore_search_checked_button()
            self.complex_articles = articles if articles else []
            self.original_complex_articles = self.complex_articles.copy()

            if not self.complex_articles:
                self.property_table.load_data([])
                if hasattr(self, 'article_label'): self.article_label.setText("선택된 단지의 매물이 없습니다.")
                self.statusBar().showMessage("선택한 단지에서 매물을 찾지 못했습니다.")
                self._update_filter_radios(self.article_filter_layout, self.article_type_group, [], 'tradeTypeName', self.filter_articles_by_trade_type)
                return

            total_articles_found = len(self.complex_articles)
            num_complexes_with_articles = len(complex_names_found)
            display_name_str = f"{original_checked_count}개 단지"
            if num_complexes_with_articles > 0:
                display_name_str = f"{complex_names_found[0]} 등 {num_complexes_with_articles}곳" if num_complexes_with_articles > 1 else complex_names_found[0]
            
            if hasattr(self, 'article_label'): self.article_label.setText(f"{display_name_str} 매물: {total_articles_found}건")
            self._update_filter_radios(self.article_filter_layout, self.article_type_group, self.original_complex_articles, 'tradeTypeName', self.filter_articles_by_trade_type)
            self.property_table.load_data(self.complex_articles)
            if hasattr(self, 'download_button'):
                self.download_button.setEnabled(True)
                self.download_button.setText(f"엑셀 다운로드 ({total_articles_found}건)")
                self.download_button.setStyleSheet(self.primary_button_style)
            self.statusBar().showMessage(f"{display_name_str}에서 {total_articles_found}개의 매물을 찾았습니다.")
        except Exception as e: 
            import traceback
            traceback.print_exc()
            self.statusBar().showMessage(f"매물 표시 오류: {e}")

    @Slot(int, bool)
    def on_complex_checkbox_toggled(self, row, checked):
        checked_count = len(self.complex_table.get_checked_items())
        if checked_count > 0:
            self.search_checked_button.setStyleSheet(self.primary_button_style + "QPushButton { border: 2px solid #2E7D32; }")
            self.search_checked_button.setText(f"선택 단지 검색({checked_count}개)")
            self.statusBar().showMessage(f"{checked_count}개 단지 선택됨. '선택 단지 검색' 버튼을 클릭하세요.")
        else:
            self.search_checked_button.setStyleSheet(self.primary_button_style)
            self.search_checked_button.setText("선택 단지 검색")
            self.statusBar().showMessage("검색할 단지를 체크박스로 선택하세요.")
        QApplication.processEvents()

    @Slot()
    def restore_search_checked_button(self):
        self.search_checked_button.setEnabled(True)
        self.search_checked_button.setText("선택 단지 검색")
        self.hide_progress_bar() 

    def show_warning_dialog(self, total_pages, total_items, complex_name):
        print(f"경고 무시: {complex_name}의 모든 {total_pages}페이지({total_items}개) 매물 자동 수집.")
        self.statusBar().showMessage(f"{complex_name}의 모든 {total_pages}페이지({total_items}개) 매물 데이터를 수집합니다...")
        return True

    @Slot(int)
    def on_article_detail_clicked(self, row):
        try:
            selected_article = None
            if hasattr(self.property_table, 'original_data') and self.property_table.original_data and 0 <= row < len(self.property_table.original_data):
                 selected_article = self.property_table.original_data[row]
            elif self.property_table.data and 0 <= row < len(self.property_table.data): 
                 selected_article = self.property_table.data[row]
            
            if not selected_article and self.complex_articles and 0 <= row < len(self.complex_articles):
                selected_article = self.complex_articles[row]
            
            if not selected_article:
                print("상세보기를 위한 매물 데이터가 없습니다.")
                return

            self.show_property_detail_from_data(row, selected_article)
        except Exception as e: 
            import traceback
            traceback.print_exc()

    def show_property_detail_from_data(self, row, property_data):
        if not property_data: 
            print(f"행 {row}의 매물 데이터가 없습니다.")
            return
        try:
            if self.detail_window_instance and self.detail_window_instance.isVisible():
                self.detail_window_instance.update_property_details(property_data)
                self.detail_window_instance.activateWindow()
                return
            self.detail_window_instance = PropertyDetailWidget(self) 
            self.detail_window_instance.update_property_details(property_data)
            self.detail_window_instance.show()
        except Exception as e: 
            import traceback
            traceback.print_exc()

    @Slot(int, bool)
    def on_article_checkbox_toggled(self, row, checked):
        checked_count = len(self.property_table.get_checked_items())
        current_total_in_table = self.property_table.rowCount()
        download_btn_text = f"엑셀 다운로드 ({current_total_in_table}건)"
        status_msg = f"총 {current_total_in_table}개 매물 표시 중."
        if checked_count > 0:
            download_btn_text = f"선택 다운로드 ({checked_count}건)"
            status_msg = f"{checked_count}개 매물 선택됨. '다운로드' 버튼으로 저장하세요."
            self.download_button.setStyleSheet(self.primary_button_style + "QPushButton { border: 2px solid #2E7D32; }")
        else: self.download_button.setStyleSheet(self.primary_button_style)
        if hasattr(self, 'download_button'): self.download_button.setText(download_btn_text)
        self.statusBar().showMessage(status_msg)
        QApplication.processEvents()

    @Slot()
    def enable_search_button(self, button_text="검색"):
        for widget in self.findChildren(QPushButton):
            if widget.text() in [button_text, "검색 중..."]: 
                 widget.setEnabled(True)
                 if button_text == "검색": widget.setText("검색")
                 break

    def create_menus(self):
        self.menubar = self.menuBar()
        file_menu = self.menubar.addMenu("파일(&F)")
        exit_action = QAction("종료(&X)", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        export_action = QAction("엑셀로 내보내기(&E)", self)
        export_action.setShortcut("Ctrl+E")
        export_action.triggered.connect(self.download_to_excel)
        file_menu.addAction(export_action)
        search_menu = self.menubar.addMenu("검색(&S)")
        keyword_search_action = QAction("키워드 검색(&K)", self)
        keyword_search_action.setShortcut("Ctrl+K")
        keyword_search_action.triggered.connect(self.fetch_from_api)
        search_menu.addAction(keyword_search_action)
        help_menu = self.menubar.addMenu("도움말(&H)")
        about_action = QAction("프로그램 정보(&A)", self)
        about_action.triggered.connect(lambda: QMessageBox.about(self, "프로그램 정보", "네이버 부동산 데이터 뷰어 v1.0.3\n\n부동산 매물 정보를 조회하고 관리합니다."))
        help_menu.addAction(about_action)

    def create_toolbar(self):
        self.toolbar = QToolBar("메인 툴바")
        self.toolbar.setMovable(False)
        self.toolbar.setIconSize(QSize(16,16))
        self.addToolBar(self.toolbar)
        self.complex_label = QLabel("단지 목록: 0개")
        self.toolbar.addWidget(self.complex_label)
        self.toolbar.addSeparator()
        self.article_label = QLabel("선택된 단지의 매물 목록: 0개")
        self.toolbar.addWidget(self.article_label)

    def _get_excel_numeric_price(self, price_str):
        if not isinstance(price_str, str) or price_str == '-': return None
        try:
            price_str_cleaned = price_str.replace(',', '').strip()
            total_won = 0
            if '억' in price_str_cleaned:
                parts = price_str_cleaned.split('억')
                eok_val_str = parts[0].strip()
                if eok_val_str and eok_val_str.replace('.', '', 1).isdigit(): total_won += float(eok_val_str) * 100000000
                man_val_str = ""
                if len(parts) > 1 and parts[1].strip(): man_val_str = parts[1].strip()
                if man_val_str and man_val_str.replace('.', '', 1).isdigit(): total_won += float(man_val_str) * 10000
            elif price_str_cleaned.replace('.', '', 1).isdigit(): total_won += float(price_str_cleaned) * 10000
            else: return price_str 
            return int(total_won)
        except (ValueError, TypeError): return price_str

    def _get_excel_numeric_rent(self, rent_str):
        if rent_str is None or rent_str == '-': return None
        if isinstance(rent_str, (int, float)): return int(rent_str * 10000)
        if isinstance(rent_str, str) and rent_str.strip().replace('.', '', 1).isdigit():
            try: return int(float(rent_str.strip()) * 10000)
            except ValueError: return rent_str 
        return rent_str 
        
    def _format_excel_date(self, date_str):
        if not date_str or not isinstance(date_str, str) or len(date_str) != 8: return date_str
        try: return f"{date_str[:4]}.{date_str[4:6]}.{date_str[6:]}"
        except: return date_str

    def download_to_excel(self):
        articles_to_process = []
        is_selection_download = False
        
        if hasattr(self, 'complex_articles') and self.complex_articles:
            checked_article_indices = self.property_table.get_checked_items()
            if checked_article_indices:
                articles_to_process = [self.complex_articles[i] for i in checked_article_indices if 0 <= i < len(self.complex_articles)]
                is_selection_download = True
                if not articles_to_process:
                     self.statusBar().showMessage("선택된 항목을 찾을 수 없습니다. 필터를 확인해주세요.")
                     return
            else:
                reply = QMessageBox.question(self, "다운로드 확인",
                    f"선택된 항목이 없습니다. 현재 목록에 있는 {len(self.complex_articles)}개의 매물을 모두 다운로드하시겠습니까?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
                if reply == QMessageBox.StandardButton.Yes: articles_to_process = self.complex_articles
                else: self.statusBar().showMessage("다운로드가 취소되었습니다."); return
        else: self.statusBar().showMessage("다운로드할 매물 데이터가 없습니다."); return

        if not articles_to_process: 
            self.statusBar().showMessage("다운로드할 데이터가 없습니다.")
            return

        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_complex_names = set(a.get('complexName', '정보없음') for a in articles_to_process if a.get('complexName'))
        base_name = "매물목록"
        if len(unique_complex_names) == 1: base_name = list(unique_complex_names)[0]
        elif len(unique_complex_names) > 1: base_name = f"{list(unique_complex_names)[0]} 등 {len(unique_complex_names)}개 단지"
        count_str = f"_{len(articles_to_process)}건"
        if is_selection_download: count_str = f"_선택항목{count_str}"
        default_filename = f"{base_name}{count_str}_{now}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "엑셀 파일 저장", default_filename, "Excel Files (*.xlsx)")
        if not file_path: return

        self.statusBar().showMessage(f"엑셀 파일 생성 중...")
        
        excel_data_rows = []
        
        excel_column_api_map = {
            '매물명': 'articleName', 
            '부동산유형': 'realEstateTypeName',
            '매물부동산유형': 'articleRealEstateTypeName',
            '거래유형': 'tradeTypeName',
            '확인유형': 'verificationTypeCode',
            '층정보': 'floorInfo',
            '가격변동상태': 'priceChangeState',
            '매매/전세가': 'dealOrWarrantPrc',
            '월세': 'rentPrc',
            '면적명': 'areaName',
            '전용면적': 'area1', # 수정됨
            '공급면적': 'area2', # 수정됨
            '방향': 'direction',
            '매물확인일': 'articleConfirmYmd', 
            '매물특징': 'articleFeatureDesc',
            '태그목록': 'tagList', 
            '동이름': 'buildingName',
            '동일주소매물수': 'sameAddrCnt',
            '동일주소최고가': 'sameAddrMaxPrc', 
            '동일주소최저가': 'sameAddrMinPrc', 
            '위도': 'latitude',
            '경도': 'longitude',
            '부동산중개사명': 'realtorName',
            '소유자확인거래': 'tradeCheckedByOwner', 
            '직거래여부': 'isDirectTrade', 
            '관심여부': 'isInterest', 
            '단지여부': 'isComplex', 
            '상세주소': 'detailAddress', 
            '안심임대인여부': 'isSafeLessorOfHug', 
            '단지명': 'complexName', 
            '단지번호': 'complexNo', 
            '주소': 'cortarAddressFromComplex',
            '지역코드': 'cortarNoFromComplex',
            '링크': 'cpPcArticleUrl', 
            '최고층': 'highFloorFromComplex',
            '최저층': 'lowFloorFromComplex',
            '최대공급면적': 'maxSupplyAreaFromComplex',
            '최대전체면적': 'maxTotalAreaFromComplex',
            '최소공급면적': 'minSupplyAreaFromComplex',
            '최소전체면적': 'minTotalAreaFromComplex',
            '총동수': 'totalDongCountFromComplex',
            '총세대수': 'totalHouseholdCountFromComplex',
            '사용승인일': 'useApproveYmdFromComplex', 
            '사용여부': 'useYnFromComplex' 
        }
        
        verification_mapping = {'NONE': '없음', 'OWNER': '소유자확인', 'REALTOR': '중개사확인', 'LESSOR': '임대인확인', 'S_VR': 'VR확인', 'SITE': '현장확인', 'NDOC1': '서류확인1', 'DOC': '서류확인', 'DOCV2': '서류확인V2', 'MOBL': '모바일확인', 'NDOC2': '서류확인2'}
        price_change_mapping = {'SAME': '변동없음', 'DOWN': '하락', 'UP': '상승', 'DECREASE': '감소', 'INCREASE': '증가', 'NEW': '신규'}

        for article_data in articles_to_process:
            excel_row = {}
            for korean_col_name in self.desired_excel_columns_korean:
                api_key = excel_column_api_map.get(korean_col_name)
                raw_value = article_data.get(api_key, '-') if api_key else '-'

                if korean_col_name == '매매/전세가':
                    excel_row[korean_col_name] = self._get_excel_numeric_price(article_data.get('dealOrWarrantPrc'))
                elif korean_col_name == '월세':
                    excel_row[korean_col_name] = self._get_excel_numeric_rent(article_data.get('rentPrc'))
                elif korean_col_name == '확인일' or korean_col_name == '사용승인일':
                    excel_row[korean_col_name] = self._format_excel_date(raw_value if isinstance(raw_value, str) else str(raw_value))
                elif korean_col_name == '태그목록':
                    excel_row[korean_col_name] = ', '.join(raw_value) if isinstance(raw_value, list) else raw_value
                elif korean_col_name == '확인유형':
                    excel_row[korean_col_name] = verification_mapping.get(str(raw_value).upper(), raw_value)
                elif korean_col_name == '가격변동상태':
                    excel_row[korean_col_name] = price_change_mapping.get(str(raw_value).upper(), raw_value)
                elif korean_col_name in ['소유자확인거래', '직거래여부', '관심여부', '단지여부', '안심임대인여부', '사용여부']:
                    excel_row[korean_col_name] = 'Y' if raw_value is True or str(raw_value).upper() == 'Y' else ('N' if raw_value is False or str(raw_value).upper() == 'N' else raw_value)
                elif korean_col_name == '주소':
                    addr_str = article_data.get('cortarAddressFromComplex', '') 
                    bldg_name = article_data.get('buildingName', '')
                    detail_addr = article_data.get('detailAddress', '')
                    
                    full_addr_parts = []
                    if addr_str : full_addr_parts.append(addr_str)
                    if bldg_name : full_addr_parts.append(bldg_name)
                    if detail_addr : full_addr_parts.append(detail_addr)
                    excel_row[korean_col_name] = " ".join(full_addr_parts).strip() if full_addr_parts else '-'
                elif korean_col_name == '매물명':
                    excel_row[korean_col_name] = article_data.get('articleName', article_data.get('complexName', '-'))
                else:
                    excel_row[korean_col_name] = raw_value if raw_value is not None else "-"
            
            excel_data_rows.append(excel_row)
            
        df = pd.DataFrame(excel_data_rows)
        df = df[self.desired_excel_columns_korean]

        try:
            df.to_excel(file_path, index=False, engine='openpyxl')
            self.statusBar().showMessage(f"엑셀 파일 저장 완료: {file_path} ({len(df)}건)")
        except Exception as e:
            self.statusBar().showMessage(f"엑셀 파일 저장 오류: {e}")
            QMessageBox.critical(self, "엑셀 저장 오류", f"엑셀 파일 저장 중 오류가 발생했습니다:\n{str(e)}")
            import traceback; traceback.print_exc()


    @Slot()
    def filter_articles_by_trade_type(self, button=None):
        if not button or not hasattr(self, 'original_complex_articles'): 
            if hasattr(self, 'property_table'): self.property_table.load_data([])
            if hasattr(self, 'article_label'): self.article_label.setText("매물 목록: 0건")
            self._update_filter_radios(self.article_filter_layout, self.article_type_group, [], 'tradeTypeName', self.filter_articles_by_trade_type)
            return

        selected_type = button.text()
        if selected_type == "전체": self.complex_articles = self.original_complex_articles.copy()
        else: self.complex_articles = [a for a in self.original_complex_articles if a.get('tradeTypeName') == selected_type]
        self.property_table.load_data(self.complex_articles)
        total_articles = len(self.complex_articles)
        if hasattr(self, 'article_label'):
            active_complex_names = list(set(art.get('complexName', '') for art in self.original_complex_articles if art.get('complexName'))) 
            if not active_complex_names and self.complex_articles:
                 active_complex_names = list(set(art.get('complexName', '') for art in self.complex_articles if art.get('complexName')))
            num_active_complexes = len(active_complex_names)
            name_prefix = "선택 단지" 
            if num_active_complexes == 1: name_prefix = active_complex_names[0]
            elif num_active_complexes > 1: name_prefix = f"{active_complex_names[0]} 등 {num_active_complexes}곳"
            self.article_label.setText(f"{name_prefix} ({selected_type}) 매물: {total_articles}건")
        self.statusBar().showMessage(f"'{selected_type}' 필터 적용: {total_articles}개 매물 표시")
        if hasattr(self, 'download_button'):
            self.download_button.setText(f"엑셀 다운로드 ({total_articles}건)")
            self.on_article_checkbox_toggled(-1, False) 

    @Slot()
    def filter_complexes_by_type(self, button=None):
        if not button or not hasattr(self, 'original_articles'): 
            if hasattr(self, 'complex_table'): self.complex_table.load_data([])
            if hasattr(self, 'complex_label'): self.complex_label.setText("단지 목록: 0개")
            self._update_filter_radios(self.complex_filter_layout, self.complex_type_group, [], 'realEstateTypeName', self.filter_complexes_by_type)
            return
        selected_type_text = button.text()
        if selected_type_text == "전체": self.articles = self.original_articles.copy()
        else: self.articles = [c for c in self.original_articles if c.get('realEstateTypeName', '') == selected_type_text]
        self.complex_table.load_data(self.articles)
        if hasattr(self, 'complex_label'): self.complex_label.setText(f"단지 목록: {len(self.articles)}개")
        self.statusBar().showMessage(f"'{selected_type_text}' 필터 적용: {len(self.articles)}개 단지 표시")